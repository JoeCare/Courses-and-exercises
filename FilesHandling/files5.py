import os, shutil, openpyxl as xl
from openpyxl.utils import get_column_letter as g_col_letter
from openpyxl.utils import column_index_from_string as col_ind_f_str
import re
"""Skrypt przyjmuje ścieżkę do pliku .xlsx zawierającego listę obecności, a następnie przetwarza ją 
uwzględniając możliwości niezarejesrowane jeszcze oficjalnie.
Potrzebne zmienne:
workbook.xlsx path with os.path.sep for Windows;
ws1['A5':'J5']

"""
# def work_sheet(wbpath,)
inputFile = 'eka'
shutil.copy('/home/morfina/Downloads/eka.xlsx','./Excel')
wb1 = xl.load_workbook('./Excel/eka.xlsx')
print(wb1.get_sheet_names())
ws1 = wb1.get_sheet_by_name('Arkusz6')
print(g_col_letter(ws1.max_column))
print(col_ind_f_str('J'))

print(ws1['A1':'J10'])
regXname = re.compile(r'W\w+ G\w+')
timeTable = []
row = []
for rowOfCellObjects in ws1['A5':'J30']:
    for cellObj in rowOfCellObjects:
        if type(cellObj.value) == str:
            searchName = regXname.search(cellObj.value)
            if searchName:
                cellObj.value = regXname.sub("John Doe", cellObj.value)
        elif type(cellObj.value) != str:
            cellObj.value = str(cellObj.value)
        # print(cellObj.coordinate, cellObj.value)
        celVal = f'{cellObj.coordinate} {cellObj.value}'
        if len(row) < 9:
            row.append(celVal)
        else:
            timeTable.append(row)
            row = []
    # print('---ROW END---')
print("Timetable", timeTable)

# for row in timeTable:
#     for cell in row:
#         if 'bilans' in cell:
#             print(row[cell])


# print(os.getcwd())
# wb2 = xl.Workbook()
# wb2s = wb2.active
#
# for row in timeTable:
#     wb2s.append(row)
#
#
# wb2.save('output1.xlsx')
#
# for row in timeTable:
#     print(row)
#     row[1] = row[1].rstrip("00:00").rstrip(" ")
#     print(row[1], row[3:9])
#     row = row[1], row[3:9]
#     timeTable1 = timeTable.append(row)
# print(timeTable1)
#
#
#
# # The outer for loop goes over each row in the slice.
# # Then, for each row, the nested for loop goes through each cell in that row.
#
#
#
#
#
#
#
#
#
#
#
#
# # with open('./Excel/readCensusExcel.py','w+') as f:
# #     f.write("#! python3 \n# "
# #             "readCensusExcel.py - Tabulates population and number of census tracts for \n#"
# #             " each county.\nExercise from Al Sweigart.")
# # maybe some day, for now its incredibly booring
#
