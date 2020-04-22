import os, shutil, openpyxl as xl
from openpyxl.utils import get_column_letter as g_col_letter
from openpyxl.utils import column_index_from_string as col_ind_f_str
import re
"""
Skrypt przyjmuje ścieżkę do pliku .xlsx zawierającego listę obecności, a następnie przetwarza ją 
uwzględniając możliwości niezarejesrowane jeszcze oficjalnie.
Potrzebne zmienne:
workbook.xlsx path with os.path.sep for Windows;
ws1['A5':'J5']
"""
# UNCOMMENT THIS:
# inputName = input("xlsManager may create changes on its current folder. "
#                  "Please type the your xlsx file name here."
#                  "or press enter to abort.")+".xlsx"
# if inputName:

inputWB = xl.load_workbook('eka.xlsx')
# this line is for testing purpose

# UNCOMMENT THIS:
# try:
#   inputWB = xl.load_workbook(inputName)
# except ?:
#   print("Supported formats are: .xlsx,.xlsm,.xltx,.xltm")

# Your workbook

inputWS = inputWB.get_sheet_by_name("Arkusz6")
# Your worksheet

table = []
row = []

for rows in inputWS.iter_rows():
    for cells in rows:
        if cells.value is not None:  # and len(str(cells.value)) >= 1:
            # print(cells.value)
            row.append(str(cells.value))
            # print(row)
    if row:
        table.append(row)
        row = []
# rewriting tuples to lists


def col_num_from_cell(value):
    """function searches for column name '' lost inside cells and returns its number"""
    for rows in table:
        for cells in rows:
            if value == cells:
                return rows.index(cells)

def sum_col_after_cell(value):
    """function takes an int indicating column position and sums its values"""
    colSum = []
    for rows in table:
        # print(rows)
        if len(rows) >= value and not rows[value].isalpha():
            colSum.append(float(rows[value]))
    return f'Sum of {value}th column values is {sum(colSum)}.'


def updated_wb():
    for rows in table:
        maxCol, maxRow = len(table), len(rows)
    table.append(tableEnd)
    print(len(table))
    return table

# m_col = inputWS.max_column
# m_row = inputWS.max_row+1
# coord = f'{str(g_col_letter(m_col))}{str(m_row)}'


sumOfColumn = [sum_col_after_cell(col_num_from_cell('bilans'))]

tableEnd = ['']*9
tableEnd[8] = sumOfColumn

# print(updated_wb())

output_wb = xl.Workbook()
output_ws = output_wb.active
for rows in updated_wb():
    #output_ws.append(rows)
    print(rows)
# rewriting list to xlsx
# outputWB.save('outputWB.xlsx')
# # and saving file


#for rows in inputWS.iter_rows(values_only=True):
#     for cells in rows:
#         if cells is not None:
#             # print(cells)
#             row.append(str(cells))
#     table.append(row)
#     row = []
# # rewriting tuples to lists
# for row in table:
#     print(row)

#
# for col in inputWS.iter_cols(values_only=True):
#     for rows in col:
#         row.append(rows)
#     print(row)





# for rows in table:
#     for cells in rows:
#         if rows[cells] == 0:
#             table.remove(rows)
#
# print(table)
#     for cells in rows:
#         print(cells)
#
# try:
#     print(absenceInd)
# except NameError as err:
#     print("No 'absencja' found in columns.")

updatewb = updated_wb()
# UNCOMMENT THIS:
outputWB = xl.Workbook()
outputWS = outputWB.active
for rows in updatewb:
    outputWS.append(rows)
# rewriting list to xlsx
outputWB.save('outputWB.xlsx')
# and saving file


