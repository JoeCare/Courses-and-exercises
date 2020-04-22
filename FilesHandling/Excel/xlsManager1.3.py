# MODULES:
import os, shutil, openpyxl as xl
from openpyxl.utils import get_column_letter as g_col_letter
from openpyxl.utils import column_index_from_string as col_ind_f_str
import re


# FUNCTIONS:
# def get_file():


def get_sheet():
    if len(inputWB.get_sheet_names()) > 1:
        print(f'Your file has more than one sheet:')
        for sheets in inputWB.get_sheet_names():
            print(sheets)
        inputSName = input("Please type the name of actual worksheet: ")
        if inputSName in inputWB.get_sheet_names():
            return inputWB.get_sheet_by_name(inputSName)
        else:
            print(f'No such sheet name. Opening last active sheet: "{inputWB.active.title}"')
            return inputWB.active
    else:
        return inputWB.active


def col_num_from_cell(table, cel_val='bilans'):
    """function searches for column name '' lost inside cells and returns its number"""
    for rows in table:
        for cells in rows:
            if cel_val == cells:
                colnum = rows.index(cells)
                return colnum


def sum_col_after_cell(table, col_num):
    """function takes an int indicating column position and sums its values"""
    if isinstance(col_num, int):
        colSum = []
        for rows in table:
            # print(rows)
            if len(rows) >= col_num and not rows[col_num].isalpha():
                colSum.append(float(rows[col_num]))
        return f'Sum of {col_num+1}th column values is {sum(colSum)}.'
    else:
        return f'Can\'t return value. Column not found.'


def updated_mysheet(table, col_num):
    maxCol = len(table) - 1
    maxRow = len(table[maxCol - 1])
    # for rows in table:
    #     maxCol, maxRow = len(table), len(rows)
    insrtion = [''] * maxRow
    insrtion[col_num] = final_sum
    table.append(insrtion)  # .append(tableEnd)
    print(maxCol, maxRow)
    return table


# PROCEDURE:
"""
Script takes the Timetable.xlsx file containing time table of single employee of some company and sums
up bilance of working time. Put the xlsManager.py file into directory containing Your .xlsx file.
"""
# inputFileName = input("xlsManager will now run through indiced path looking "
#                       "for file named Timetable and parse it. Type the"
#                       "Press enter to abort.")+".xlsx"

# if inputFileName:

inputWB = xl.load_workbook('eka.xlsx')

inputWS = get_sheet()
my_sheet = []
row = []

for rows in inputWS.iter_rows():
    for cells in rows:
        if cells.value is not None:  # and len(str(cells.value)) >= 1:
            # print(cells.value)
            row.append(str(cells.value))
            # print(row)
    if row:
        my_sheet.append(row)
        row = []
    # print(rows)

column_number = col_num_from_cell(my_sheet)
final_sum = sum_col_after_cell(my_sheet, column_number)

output_wb = xl.Workbook()
output_ws = output_wb.active
# creating new workbook with worksheet
for rows in updated_mysheet(my_sheet, column_number):
    output_ws.append(rows)
    print(rows)
# putting lists back to openxl tuples
output_wb.save('outputWB2.xlsx')
# and saving to the file

# for file in os.listdir():
#     regFile = re.compile(r'^outputWB.\.xlsx')
#     if regFile.search(file):
#         os.remove(file)
#         print('Deleted file: ', file)
#     else:
#         print('Left file: ', file)
# little cleanup

