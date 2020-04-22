import os, shutil, openpyxl as xl
from openpyxl.utils import get_column_letter as g_col_letter
from openpyxl.utils import column_index_from_string as col_ind_f_str
import re

"""
Skrypt przyjmuje ścieżkę do pliku .xlsx zawierającego listę obecności, a następnie przetwarza ją 
uwzględniając możliwości niezarejesrowane jeszcze oficjalnie.
Potrzebne zmienne:
workbook.xlsx path with os.path.sep for Windows;
ws1['A5':'J5']
"""
# UNCOMMENT THIS:
# inputName = input("xlsManager may create changes on its current folder. "
#                  "Please type the your xlsx file name here."
#                  "or press enter to abort.")+".xlsx"
# if inputName:

inputWB = xl.load_workbook('eka.xlsx')
# this line is for testing purpose

# UNCOMMENT THIS:
# try:
#   inputWB = xl.load_workbook(inputName)
# except ?:
#   print("Supported formats are: .xlsx,.xlsm,.xltx,.xltm")

# Your workbook

if len(inputWB.get_sheet_names()) > 1:
    print(f'Your file has more than one sheet:\n')
    for sheets in inputWB.get_sheet_names():
        print(sheets)
    inputWS = input("Please type the name of actual worksheet: ")
else:
    inputWS = inputWB.active

inputWS = inputWB.get_sheet_by_name("Arkusz6")
# Your worksheet

table = []
row = []

for rows in inputWS.iter_rows():
    for cells in rows:
        if cells.value is not None:  # and len(str(cells.value)) >= 1:
            # print(cells.value)
            row.append(str(cells.value))
            # print(row)
    if row:
        table.append(row)
        row = []


# rewriting tuples to lists


def col_num_from_cell(cel_val='bilans'):
    """function searches for column name '' lost inside cells and returns its number"""
    for rows in table:
        for cells in rows:
            if cel_val == cells:
                colnum = rows.index(cells)
                return colnum


def sum_col_after_cell(col_num):
    """function takes an int indicating column position and sums its values"""
    if isinstance(col_num, int):
        colSum = []
        for rows in table:
            # print(rows)
            if len(rows) >= col_num and not rows[col_num].isalpha():
                colSum.append(float(rows[col_num]))
        return f'Sum of {col_num}th column values is {sum(colSum)}.'
    else:
        return f'Can\'t return value. Column not found.'


def updated_wb():
    maxCol = len(table) - 1
    maxRow = len(table[maxCol - 1])
    # for rows in table:
    #     maxCol, maxRow = len(table), len(rows)
    insrtion = [''] * maxRow
    insrtion[len(insrtion) - 1] = finalSumFromColumn
    table.append(insrtion)  # .append(tableEnd)
    print(maxCol, maxRow)
    return table


# m_col = inputWS.max_column
# m_row = inputWS.max_row+1
# coord = f'{str(g_col_letter(m_col))}{str(m_row)}'


finalSumFromColumn = sum_col_after_cell(col_num_from_cell())  # inputColVal
#
tableEnd = [''] * 9
tableEnd[8] = finalSumFromColumn

# searching for coords of last cell under the table to put the summary string into it

for file in os.listdir():
    regFile = re.compile(r'outputWB.\.xlsx')
    if regFile.search(file):
        print('Deleted files: ', file)
    else:
        print('Left files: ', file)

output_wb = xl.Workbook()
output_ws = output_wb.active
for rows in updated_wb():
    output_ws.append(rows)
    print(rows)
# rewriting list to xlsx
output_wb.save('outputWB2.xlsx')
# # # and saving file

#
# updatewb = updated_wb()
# # UNCOMMENT THIS:
# outputWB = xl.Workbook()
# outputWS = outputWB.active
# for rows in updatewb:
#     # outputWS.append(rows)
#     print(rows)
# rewriting list to xlsx
# outputWB2.save('outputWB.xlsx')
# and saving file
