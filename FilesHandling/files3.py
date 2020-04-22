import openpyxl
from openpyxl.utils import get_column_letter
import os
# many many unsuccesful trials in that part, but... everybody learns at their own tempo... Aren't they? : o

# first I had to install openpyxl into virtual env by typing into Terminal console
# cd ~/PycharmProjects/Project1
# pip3 install openpyxl
# (i also got the 'Cache entry deserialization failed, entry ignored' msg before installation,
# but then it's been Successfully installed. Anybody know what that warn means?
# I could also just install package into default python3 dir but... I didn't knew how to make PyCharm use it
# rather her own venv xd
# So in those two cases I'm open to suggestions.

os.chdir('/home/morfina/Downloads/')
exc1 = openpyxl.load_workbook('exc1.xlsx')
# hm... now it tells me that there is ModuleNotFoundError: No module named 'openpyxl'
# so I uninstall what I tough I installed through ubuntu Terminal and try to install by PyCharm Terminal
# and now it seems to work

print(exc1.get_sheet_names())
# returns the list of sheets
sheet3 = exc1.get_sheet_by_name('Arkusz3')
print(sheet3)
# >>> <Worksheet "Arkusz3"> - so now we've got 3rd sheet declared under a variable
print(type(sheet3.title), sheet3.title)
# >>> <class 'str'> Arkusz3
print(type(exc1.active), exc1.active)
# >>> <class 'openpyxl.worksheet.worksheet.Worksheet'> <Worksheet "Arkusz1"> - last used worksheet and class info

print(type(sheet3['A5']), sheet3['A5'], sheet3['A5'].value, sheet3['A5'].row, sheet3['A5'].column, sheet3['A5'].coordinate)
# >>> <class 'openpyxl.cell.cell.Cell'> <Cell 'Arkusz3'.A5> wejście 5 1 A5
# bunch of usable info about variable holding particular Cell:
print(f"Here we've got \"{sheet3['A5']}\" of a type \"{type(sheet3['A5'])}\" with a value of \"{sheet3['A5'].value}\" "
      f"in it. \nIt is placed in the \"{sheet3['A5'].row}\" row and \"{sheet3['A5'].column}\" column so its "
      f"coordinates are: \"{sheet3['A5'].coordinate}\"")
print(sheet3.cell(row=5, column=1))
# we can access cells similar to pandas by bracket notation [] as well as method .cell() which however may be
# more reliable because it uses only integers to indicate cell's position in the worksheet
# notation based on letters may be tricky after 26th column, because once we've got no more letters in the alphabeth
# columns are signed with two or more letters:
print(sheet3['Z1'].column, sheet3['ZA1'].column, sheet3['DD1'].column)
# as Z is perceived as 26th letter ZA means that we passed whole alphabeth 26 times (677/26=26) so
# ZA is 677th column. Another A for ['ZAA1'] require to multiplicate 677 another 26 times... Just for curiosity
# one more example: DD1 will be 108th column, becouse it is 4 times 26 and plus another 4 :)

col = sheet3['C5':'C81']

tuple = ('s','a','k')
# print(sheet3.values)
# for v in sheet3.values:
#       print(v)
absences = []
afterhours = []

print(len(sheet3["C"]))
for row in sheet3["C"]:
      cell = str(row.value)
      if "-" in cell:
            absences.append(float(cell))
      elif "-" not in cell:
            if not cell.isalpha():
                  afterhours.append(float(cell))
print(afterhours, sum(afterhours))
print(absences, sum(absences))
print(len(absences),len(afterhours))
print(sum(afterhours)+sum(absences)+(4*8))
#
# sheet = ' '.join(sheet3["C5:C81"])
# print("Sum:",sheet)
# print(allValues)
# floats = []
# negative = []
# positive = []
# for row in sheet3["C"]:
#       if not str(row.value).isalpha():
#             floats.append(float(row.value))
#             print(float(row.value))
# print(floats.
#       # if "-" in cell:
#       #       absences.append(float(cell))
#       # elif "-" not in cell:
#       #       if not cell.isalpha():
#       #             afterhours.append(float(cell))
# print(sum(absences)+sum(afterhours))

tupl = sheet3["C"]
print(tupl)
col_list = list(tupl)
print(col_list)
columnC = []
columnN = []
for cell in col_list:
      cell1 = cell.value
      if type(cell1) == int or type(cell1) == float and cell1 > 0:
                  columnC.append(cell1)
                  columnC.sort(reverse=True)
      if type(cell1) == int or type(cell1) == float:
            columnN.append(cell1)
            columnN.sort(reverse=True)

print(columnC,'\n',columnN)
columnCpos = ''.join(map(str,columnC))
columnCneg = ''.join(map(str,columnN))
print(sum(columnCpos))
print(sum(columnCneg))
#
# for cell in tupl:
#       value = cell.value
#       print(type(value))

# floats = []
# ints = []
# for cell in lis:
#       value = cell.value
#       # print(type(value))
#       if type(value) == int:
#             ints.append(value)
#       elif type(value) == float:
#             floats.append(value)
# print(ints,'\n',floats)
# ints.remove(0)
# print(ints)
